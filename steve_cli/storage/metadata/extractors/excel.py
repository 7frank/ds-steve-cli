from __future__ import annotations

import io

from steve_cli.storage.metadata.port import ColumnMetadata, FileMetadata, MetadataExtractorPort


class ExcelExtractor(MetadataExtractorPort):
    extensions = (".xlsx", ".xls", ".xlsm")

    def extract(self, data: bytes, path: str) -> FileMetadata:
        try:
            import openpyxl
        except ImportError as exc:
            raise ImportError(
                "openpyxl is required for Excel metadata extraction. "
                "Install it with: pip install steve-cli[excel]"
            ) from exc

        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        sheet_names = wb.sheetnames
        ws = wb.active

        rows_iter = ws.iter_rows(values_only=True)
        header_row = next(rows_iter, None)
        columns = (
            [ColumnMetadata(name=str(h) if h is not None else f"col_{i}", type="unknown")
             for i, h in enumerate(header_row)]
            if header_row else []
        )

        row_count = sum(1 for _ in rows_iter)
        wb.close()

        return FileMetadata(
            format="excel",
            size_bytes=len(data),
            rows=row_count,
            columns=columns,
            extra={
                "sheets": sheet_names,
                "sheet_count": len(sheet_names),
            },
        )
